#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
亚马逊 FBA 入库货件 —— 交互式数据生成脚本
============================================
支持自定义选择字段、自定义字段值、控制数据条数，交互式造数据。

功能：
  - 交互式选择要生成的字段
  - 每个字段可选：自动生成 / 固定值 / 自定义范围 / 留空
  - 支持 JSON / Excel 两种模板格式，保存/加载配置，下次直接复用
  - 三种输出方式：CSV / SQL / 直连MySQL
  - 数据库增强：自动建表、清空表、连接测试、dry-run、配置文件

用法：
  # 交互式模式（推荐）
  python generate_inbound_data.py

  # 导出空白 Excel 模板，用 Excel 编辑后使用
  python generate_inbound_data.py --export-excel-template

  # 使用 Excel 模板生成数据
  python generate_inbound_data.py --template-excel my_config.xlsx --count 50 --output csv

  # 使用 JSON 模板快速生成
  python generate_inbound_data.py --template default

  # 命令行模式（非交互式，全部自动）
  python generate_inbound_data.py --count 50 --output csv

  # 直连 MySQL 写入（自动建表 + 清空旧数据）
  python generate_inbound_data.py --count 50 --output db --create-table --truncate

  # 仅预览 SQL 不实际写入（dry-run）
  python generate_inbound_data.py --count 10 --output db --dry-run --create-table

  # 测试数据库连接
  python generate_inbound_data.py --test-conn --host 192.168.1.100 --user admin --password xxx

  # 导出建表 DDL
  python generate_inbound_data.py --export-ddl

  # 使用 db_config.json 保存数据库连接配置
  python generate_inbound_data.py --count 50 --output db --db-config ./my_db_config.json

依赖安装：
  pip install faker pymysql openpyxl
"""

import argparse
import csv
import json
import os
import random
import sys
import time
from datetime import datetime, timedelta

try:
    from faker import Faker
except ImportError:
    print("缺少 faker 库，请执行: pip install faker")
    sys.exit(1)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ============================================================
# 字段定义 —— 每个字段的元信息
# ============================================================

SHIPMENT_FIELDS = {
    "MARKETPLACE_ID": {
        "label": "市场ID",
        "type": "choice",
        "choices": [
            "ATVPDKIKX0DER", "A2EUQ1WTGCTBG2", "A1AM78C64UM0Y8",
            "A1PA6795UKMFR9", "A1RKKUPIHCS9HS", "A13V1IB3VIYZZH",
            "APJ6JRA9NG5V4", "A1F83G8C2ARO7P", "A21TJRUUN4KGV",
            "A39IBJ37TRP1C6", "A1VC38T7YXB528",
        ],
        "required": True,
        "default_gen": "random_choice",
    },
    "SELLER_ID": {
        "label": "卖家ID",
        "type": "string",
        "pattern": "A + 13位字母数字",
        "required": True,
        "default_gen": "seller_id",
    },
    "SHIPMENT_ID": {
        "label": "货件ID",
        "type": "string",
        "pattern": "FBA + 6~10位字母数字",
        "required": True,
        "default_gen": "shipment_id",
    },
    "SHIPMENT_NAME": {
        "label": "货件名称",
        "type": "string",
        "default_gen": "shipment_name",
    },
    "SHIP_FROM_NAME": {
        "label": "发货方名称",
        "type": "string",
        "default_gen": "company",
    },
    "SHIP_FROM_ADDRESS_LINE1": {
        "label": "地址1",
        "type": "string",
        "default_gen": "street_address",
    },
    "SHIP_FROM_ADDRESS_LINE2": {
        "label": "地址2",
        "type": "string",
        "default_gen": "secondary_address",
        "nullable": True,
    },
    "SHIP_FROM_DISTRICT_OR_COUNTY": {
        "label": "区/县",
        "type": "string",
        "default_gen": "district",
        "nullable": True,
    },
    "SHIP_FROM_CITY": {
        "label": "城市",
        "type": "string",
        "default_gen": "city",
    },
    "SHIP_FROM_STATE_OR_PROVINCE_CODE": {
        "label": "州/省",
        "type": "string",
        "default_gen": "state",
    },
    "SHIP_FROM_POSTAL_CODE": {
        "label": "邮政编码",
        "type": "string",
        "default_gen": "zipcode",
    },
    "SHIP_FROM_COUNTRY_CODE": {
        "label": "国家代码",
        "type": "choice",
        "choices": ["US", "CA", "MX", "DE", "ES", "FR", "IT", "GB", "IN", "AU", "JP"],
        "default_gen": "random_choice",
    },
    "DESTINATION_FULFILLMENT_CENTER_ID": {
        "label": "亚马逊配送中心ID",
        "type": "string",
        "default_gen": "fc_id",
    },
    "LABEL_PREP_TYPE": {
        "label": "标签类型",
        "type": "choice",
        "choices": ["NO_LABEL", "SELLER_LABEL", "AMAZON_LABEL"],
        "default_gen": "random_choice",
    },
    "SHIPMENT_STATUS": {
        "label": "货件状态",
        "type": "choice",
        "choices": {
            "1": "处理中", "2": "已发货", "3": "在途", "4": "已交付",
            "5": "检查中", "6": "接收中", "7": "已完成", "8": "已取消",
            "9": "已删除", "10": "错误",
        },
        "default_gen": "weighted_status",
    },
    "ARE_CASES_REQUIRED": {
        "label": "是否原厂包装",
        "type": "choice",
        "choices": ["True", "False"],
        "default_gen": "random_bool",
    },
    "CONFIRMED_NEED_BY_DATE": {
        "label": "确认需求日期",
        "type": "string",
        "default_gen": "need_by_date",
        "nullable": True,
        "depends_on": "SHIPMENT_STATUS",
    },
    "BOX_CONTENTS_SOURCE": {
        "label": "箱内容来源",
        "type": "choice",
        "choices": ["NONE", "FBA_PACKAGING", "2D_BARCODE"],
        "default_gen": "random_choice",
    },
    "TOTAL_UNITS": {
        "label": "总数量",
        "type": "int",
        "range": [10, 5000],
        "default_gen": "random_int",
    },
    "CURRENCY_CODE": {
        "label": "货币代码",
        "type": "choice",
        "choices": ["USD", "CAD", "EUR", "GBP", "JPY", "INR", "AUD", "MXN"],
        "default_gen": "random_choice",
    },
    "FEE_PER_UNIT": {
        "label": "单价费用",
        "type": "float",
        "range": [0.1, 5.0],
        "default_gen": "random_float",
    },
    "TOTAL_FEE": {
        "label": "总费用",
        "type": "float",
        "default_gen": "computed_total_fee",
        "auto": True,
    },
    "AMAZON_REFERENCE_ID": {
        "label": "预约号",
        "type": "string",
        "default_gen": "reference_id",
        "nullable": True,
    },
    "CREATE_TIME": {
        "label": "创建时间",
        "type": "datetime",
        "default_gen": "create_time",
        "auto": True,
    },
    "UPDATE_TIME": {
        "label": "更新时间",
        "type": "datetime",
        "default_gen": "update_time",
        "auto": True,
        "nullable": True,
    },
}

ITEM_FIELDS = {
    "MARKETPLACE_ID": {
        "label": "市场ID",
        "type": "inherit",
        "inherit_from": "shipment",
        "required": True,
    },
    "SELLER_ID": {
        "label": "卖家ID",
        "type": "inherit",
        "inherit_from": "shipment",
        "required": True,
    },
    "SHIPMENT_ID": {
        "label": "货件ID",
        "type": "inherit",
        "inherit_from": "shipment",
        "required": True,
    },
    "SELLER_SKU": {
        "label": "卖家SKU",
        "type": "string",
        "pattern": "自定义格式",
        "required": True,
        "default_gen": "seller_sku",
    },
    "FNSKU": {
        "label": "FNSKU",
        "type": "string",
        "pattern": "X00 + 7位字母数字",
        "default_gen": "fnsku",
    },
    "QUANTITY_SHIPPED": {
        "label": "发货数量",
        "type": "int",
        "range": [1, 500],
        "default_gen": "quantity_shipped",
    },
    "QUANTITY_RECEIVED": {
        "label": "接收数量",
        "type": "int",
        "default_gen": "quantity_received",
        "auto": True,
    },
    "QUANTITY_IN_CASE": {
        "label": "每箱数量",
        "type": "int",
        "range": [1, 24],
        "default_gen": "random_int_nullable",
        "nullable": True,
    },
    "CREATE_TIME": {
        "label": "创建时间",
        "type": "inherit",
        "inherit_from": "shipment",
        "auto": True,
    },
    "UPDATE_TIME": {
        "label": "更新时间",
        "type": "inherit",
        "inherit_from": "shipment",
        "auto": True,
        "nullable": True,
    },
    "ALL_SIGN_TIME": {
        "label": "全部签收时间",
        "type": "datetime",
        "default_gen": "all_sign_time",
        "auto": True,
        "nullable": True,
    },
}

# 货件状态权重
SHIPMENT_STATUS_WEIGHTS = [5, 15, 10, 10, 3, 5, 30, 10, 8, 4]

# 配置模板保存路径
TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")
EXCEL_TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_templates")


# ============================================================
# 数据生成器
# ============================================================

class InboundDataGenerator:
    """亚马逊入库货件数据生成器"""

    def __init__(self, seed=None):
        self.fake = Faker('zh_CN')
        self.fake_en = Faker('en_US')
        if seed is not None:
            Faker.seed(seed)
            random.seed(seed)

        self._used_shipment_ids = set()
        self._seller_pool = [self._gen_seller_id() for _ in range(20)]

    def _gen_seller_id(self):
        return 'A' + ''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', k=13))

    def _gen_shipment_id(self):
        while True:
            prefix = random.choice(['FBA15', 'FBA18', 'FBA19', 'FBA20', 'FBA21', 'FBA22', 'FBA23', 'FBA24'])
            suffix = ''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', k=random.randint(6, 10)))
            sid = prefix + suffix
            if sid not in self._used_shipment_ids:
                self._used_shipment_ids.add(sid)
                return sid

    def _gen_seller_sku(self):
        patterns = [
            lambda: f"SKU-{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', k=6))}",
            lambda: f"{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ', k=3))}-{''.join(random.choices('0123456789', k=5))}",
            lambda: f"PROD-{random.randint(10000, 99999)}",
            lambda: f"AS-{self.fake_en.bothify('??##??##')}",
        ]
        return random.choice(patterns)()

    def _gen_fnsku(self):
        return 'X00' + ''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', k=7))

    def _gen_fc_id(self):
        prefixes = ['PHX', 'LAX', 'SEA', 'OAK', 'SDF', 'IND', 'CMH', 'AVP', 'ABE', 'BWI',
                    'CLT', 'RDU', 'JAX', 'MCO', 'TPA', 'DFW', 'IAH', 'SAT', 'DEN', 'SLC']
        suffix = ''.join(random.choices('0123456789', k=random.randint(1, 2)))
        return random.choice(prefixes) + suffix

    def _random_date(self, days_back=180):
        end = datetime.now()
        start = end - timedelta(days=days_back)
        delta = int((end - start).total_seconds())
        return start + timedelta(seconds=random.randint(0, delta))

    # ---- 核心方法：根据用户配置生成单条数据 ----

    def generate_field_value(self, field_name, field_config, user_config, context=None):
        """
        根据用户配置生成单个字段的值

        user_config 结构:
            "mode": "auto" | "fixed" | "range" | "null"
            "fixed_value": "xxx"    (mode=fixed 时)
            "range_min": 1          (mode=range 时)
            "range_max": 100        (mode=range 时)
            "choices": ["A","B"]    (mode=range 时，枚举选择)
        """
        mode = user_config.get("mode", "auto")

        # mode=null：不生成，返回 None
        if mode == "null":
            return None

        # mode=fixed：使用固定值
        if mode == "fixed":
            val = user_config.get("fixed_value", "")
            return self._cast_value(val, field_config)

        # mode=range：范围内随机
        if mode == "range":
            ftype = field_config.get("type", "string")
            if ftype == "int":
                lo = user_config.get("range_min", field_config.get("range", [0, 100])[0])
                hi = user_config.get("range_max", field_config.get("range", [0, 100])[1])
                return random.randint(int(lo), int(hi))
            elif ftype == "float":
                lo = user_config.get("range_min", field_config.get("range", [0.0, 1.0])[0])
                hi = user_config.get("range_max", field_config.get("range", [0.0, 1.0])[1])
                return round(random.uniform(float(lo), float(hi)), 2)
            elif ftype == "choice":
                choices = user_config.get("choices", field_config.get("choices", []))
                if isinstance(choices, dict):
                    choices = list(choices.keys())
                return random.choice(choices) if choices else None
            elif ftype == "string":
                choices = user_config.get("choices", [])
                if choices:
                    return random.choice(choices)
                lo = user_config.get("range_min", 5)
                hi = user_config.get("range_max", 10)
                return ''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', k=random.randint(int(lo), int(hi))))

        # mode=auto：自动生成
        gen_type = field_config.get("default_gen", "")

        if gen_type == "random_choice":
            choices = field_config.get("choices", [])
            if isinstance(choices, dict):
                choices = list(choices.keys())
            return random.choice(choices) if choices else None

        elif gen_type == "seller_id":
            return random.choice(self._seller_pool)

        elif gen_type == "shipment_id":
            return self._gen_shipment_id()

        elif gen_type == "shipment_name":
            return f"Shipment-{self.fake_en.bothify('##??####')}-{random.randint(1, 999)}"

        elif gen_type == "company":
            return self.fake.company()

        elif gen_type == "street_address":
            return self.fake.street_address()

        elif gen_type == "secondary_address":
            return self.fake_en.secondary_address() if random.random() > 0.4 else None

        elif gen_type == "district":
            return self.fake.city_suffix() if random.random() > 0.5 else None

        elif gen_type == "city":
            return self.fake.city()

        elif gen_type == "state":
            return self.fake_en.state_abbr() if random.random() > 0.3 else self.fake_en.state()

        elif gen_type == "zipcode":
            return self.fake_en.zipcode()

        elif gen_type == "fc_id":
            return self._gen_fc_id()

        elif gen_type == "weighted_status":
            return random.choices(
                list(range(1, 11)),
                weights=SHIPMENT_STATUS_WEIGHTS,
                k=1
            )[0]

        elif gen_type == "random_bool":
            return random.choice([True, False])

        elif gen_type == "random_int":
            r = field_config.get("range", [1, 100])
            return random.randint(r[0], r[1])

        elif gen_type == "random_int_nullable":
            r = field_config.get("range", [1, 24])
            return random.randint(r[0], r[1]) if random.random() > 0.3 else None

        elif gen_type == "random_float":
            r = field_config.get("range", [0.1, 5.0])
            return round(random.uniform(r[0], r[1]), 2)

        elif gen_type == "need_by_date":
            status = context.get("SHIPMENT_STATUS", 1) if context else 1
            if status in [2, 3, 4, 5, 6, 7]:
                base = datetime.now()
                return (base + timedelta(days=random.randint(7, 30))).strftime('%Y-%m-%d')
            return None

        elif gen_type == "computed_total_fee":
            total_units = context.get("TOTAL_UNITS", 0) if context else 0
            fee_per_unit = context.get("FEE_PER_UNIT", 0) if context else 0
            if total_units and fee_per_unit:
                return round(fee_per_unit * total_units, 2)
            return round(random.uniform(10, 5000), 2)

        elif gen_type == "reference_id":
            status = context.get("SHIPMENT_STATUS", 1) if context else 1
            if status >= 2:
                return ''.join(random.choices('0123456789', k=10))
            return None

        elif gen_type == "create_time":
            return self._random_date().strftime('%Y-%m-%d %H:%M:%S')

        elif gen_type == "update_time":
            status = context.get("SHIPMENT_STATUS", 1) if context else 1
            create = context.get("CREATE_TIME", "") if context else ""
            if status >= 2 and create:
                ct = datetime.strptime(create, '%Y-%m-%d %H:%M:%S')
                return (ct + timedelta(hours=random.randint(1, 720))).strftime('%Y-%m-%d %H:%M:%S')
            return None

        elif gen_type == "seller_sku":
            return self._gen_seller_sku()

        elif gen_type == "fnsku":
            return self._gen_fnsku()

        elif gen_type == "quantity_shipped":
            status = context.get("SHIPMENT_STATUS", 1) if context else 1
            if status == 8:
                return 0
            return random.randint(1, 500)

        elif gen_type == "quantity_received":
            status = context.get("SHIPMENT_STATUS", 1) if context else 1
            shipped = context.get("QUANTITY_SHIPPED", 0) if context else 0
            if status in [1, 2, 3]:
                return 0
            elif status in [4, 5, 6]:
                return random.randint(int(shipped * 0.5), shipped)
            elif status == 7:
                return random.randint(int(shipped * 0.9), shipped)
            elif status == 8:
                return 0
            return random.randint(0, shipped)

        elif gen_type == "all_sign_time":
            status = context.get("SHIPMENT_STATUS", 1) if context else 1
            shipped = context.get("QUANTITY_SHIPPED", 0) if context else 0
            received = context.get("QUANTITY_RECEIVED", 0) if context else 0
            if status == 7 and received >= shipped * 0.95:
                create = context.get("CREATE_TIME", "") if context else ""
                if create:
                    ct = datetime.strptime(create, '%Y-%m-%d %H:%M:%S')
                    return (ct + timedelta(days=random.randint(3, 30))).strftime('%Y-%m-%d %H:%M:%S')
            return None

        # 兜底
        return None

    def _cast_value(self, val, field_config):
        """将字符串值转换为字段对应类型"""
        ftype = field_config.get("type", "string")
        if ftype == "int":
            return int(val)
        elif ftype == "float":
            return float(val)
        elif ftype == "choice":
            choices = field_config.get("choices", {})
            if isinstance(choices, dict):
                if val in choices:
                    return int(val) if val.isdigit() else val
            return val
        elif ftype == "datetime":
            return val
        return val

    def generate_shipment(self, field_config):
        """
        根据用户配置生成一条货件数据

        Args:
            field_config: dict, {field_name: {mode, fixed_value, range_min, range_max, choices}}

        Returns:
            dict: 一条货件数据
        """
        row = {}
        context = {}

        # 分两轮：第一轮生成普通字段，第二轮生成依赖其他字段的自动字段
        auto_fields = []
        for fname, fdef in SHIPMENT_FIELDS.items():
            if fdef.get("auto") or fdef.get("type") == "inherit":
                auto_fields.append(fname)
                continue
            cfg = field_config.get(fname, {"mode": "auto"})
            val = self.generate_field_value(fname, fdef, cfg, context)
            row[fname] = val
            context[fname] = val

        # 第二轮：生成自动/依赖字段
        for fname in auto_fields:
            fdef = SHIPMENT_FIELDS[fname]
            cfg = field_config.get(fname, {"mode": "auto"})
            val = self.generate_field_value(fname, fdef, cfg, context)
            row[fname] = val
            context[fname] = val

        return row

    def generate_items_for_shipment(self, shipment, item_field_config, min_items=1, max_items=5):
        """为一个货件生成明细数据"""
        num_items = random.randint(min_items, max_items)
        items = []
        used_skus = set()

        for _ in range(num_items):
            row = {}
            context = dict(shipment)

            for fname, fdef in ITEM_FIELDS.items():
                # 继承字段
                if fdef.get("type") == "inherit":
                    row[fname] = shipment.get(fname)
                    context[fname] = shipment.get(fname)
                    continue

                cfg = item_field_config.get(fname, {"mode": "auto"})

                # SKU 需要去重
                if fname == "SELLER_SKU":
                    while True:
                        val = self.generate_field_value(fname, fdef, cfg, context)
                        if val not in used_skus:
                            used_skus.add(val)
                            break
                    row[fname] = val
                    context[fname] = val
                else:
                    val = self.generate_field_value(fname, fdef, cfg, context)
                    row[fname] = val
                    context[fname] = val

            items.append(row)

        return items

    def generate_batch(self, count, shipment_config, item_config, min_items=1, max_items=5):
        """批量生成数据"""
        shipments = []
        items = []

        for i in range(count):
            s = self.generate_shipment(shipment_config)
            s_items = self.generate_items_for_shipment(s, item_config, min_items, max_items)
            shipments.append(s)
            items.extend(s_items)

            if (i + 1) % 100 == 0:
                print(f"  已生成 {i + 1}/{count} 个货件...")

        return shipments, items


# ============================================================
# 交互式配置
# ============================================================

def print_separator(char="=", length=60):
    print(char * length)


def print_title(title):
    print_separator()
    print(f"  {title}")
    print_separator()


def input_with_default(prompt, default=""):
    val = input(f"{prompt} [{default}]: ").strip()
    return val if val else default


def yes_no(prompt, default="n"):
    val = input(f"{prompt} (y/n) [{default}]: ").strip().lower()
    return val in ("y", "yes") if val else default in ("y", "yes")


def configure_field_interactive(field_name, field_def):
    """
    交互式配置单个字段

    Returns:
        dict: {mode, fixed_value, range_min, range_max, choices}
    """
    label = field_def.get("label", field_name)
    ftype = field_def.get("type", "string")
    nullable = field_def.get("nullable", False)
    required = field_def.get("required", False)

    print(f"\n  📌 {field_name} ({label}) — 类型: {ftype}")

    # 显示可选项
    if ftype == "choice":
        choices = field_def.get("choices", {})
        if isinstance(choices, dict):
            print(f"     可选值:")
            for k, v in choices.items():
                print(f"       {k}: {v}")
        elif isinstance(choices, list):
            print(f"     可选值: {', '.join(str(c) for c in choices)}")

    if ftype in ("int", "float"):
        rng = field_def.get("range", [])
        if rng:
            print(f"     默认范围: {rng[0]} ~ {rng[1]}")

    pattern = field_def.get("pattern", "")
    if pattern:
        print(f"     格式: {pattern}")

    # 选择模式
    if nullable and not required:
        print(f"  生成模式:")
        print(f"    1. 自动生成（随机）")
        print(f"    2. 固定值（每条数据都一样）")
        print(f"    3. 自定义范围/选项")
        print(f"    4. 留空（NULL）")
        mode_choice = input_with_default("  请选择", "1")
    else:
        print(f"  生成模式:")
        print(f"    1. 自动生成（随机）")
        print(f"    2. 固定值（每条数据都一样）")
        print(f"    3. 自定义范围/选项")
        mode_choice = input_with_default("  请选择", "1")

    config = {}

    if mode_choice == "4" and nullable:
        config["mode"] = "null"
        print(f"  ✅ {field_name} → 留空(NULL)")
        return config

    elif mode_choice == "2":
        # 固定值
        config["mode"] = "fixed"
        if ftype == "choice":
            choices = field_def.get("choices", {})
            if isinstance(choices, dict):
                val = input_with_default(f"  请输入值 ({'/'.join(choices.keys())})",
                                         list(choices.keys())[0] if choices else "")
            else:
                val = input_with_default(f"  请输入值 ({'/'.join(str(c) for c in choices)})",
                                         str(choices[0]) if choices else "")
        else:
            val = input_with_default("  请输入固定值", "")
        config["fixed_value"] = val
        print(f"  ✅ {field_name} → 固定值: {val}")

    elif mode_choice == "3":
        # 自定义范围/选项
        config["mode"] = "range"
        if ftype == "int":
            lo = input_with_default("  最小值", str(field_def.get("range", [0, 100])[0]))
            hi = input_with_default("  最大值", str(field_def.get("range", [0, 100])[1]))
            config["range_min"] = int(lo)
            config["range_max"] = int(hi)
            print(f"  ✅ {field_name} → 范围: {lo} ~ {hi}")
        elif ftype == "float":
            lo = input_with_default("  最小值", str(field_def.get("range", [0.0, 1.0])[0]))
            hi = input_with_default("  最大值", str(field_def.get("range", [0.0, 1.0])[1]))
            config["range_min"] = float(lo)
            config["range_max"] = float(hi)
            print(f"  ✅ {field_name} → 范围: {lo} ~ {hi}")
        elif ftype == "choice":
            choices = field_def.get("choices", {})
            if isinstance(choices, dict):
                print(f"  可选值: {', '.join(f'{k}({v})' for k, v in choices.items())}")
                selected = input_with_default("  输入要选的值(多个用逗号分隔)", ",".join(list(choices.keys())[:3]))
                config["choices"] = [c.strip() for c in selected.split(",")]
            else:
                selected = input_with_default(f"  输入要选的值(多个用逗号分隔)", ",".join(str(c) for c in choices[:3]))
                config["choices"] = [c.strip() for c in selected.split(",")]
            print(f"  ✅ {field_name} → 从以下选项中随机: {config['choices']}")
        elif ftype == "string":
            print(f"  字符串自定义模式:")
            print(f"    a. 从几个固定值中随机选")
            print(f"    b. 指定长度范围随机生成")
            sub = input_with_default("  请选择", "a")
            if sub == "a":
                vals = input("  输入可选值(多个用逗号分隔): ").strip()
                config["choices"] = [v.strip() for v in vals.split(",")]
                print(f"  ✅ {field_name} → 从以下值中随机: {config['choices']}")
            else:
                lo = input_with_default("  最短长度", "5")
                hi = input_with_default("  最长长度", "10")
                config["range_min"] = int(lo)
                config["range_max"] = int(hi)
                print(f"  ✅ {field_name} → 长度范围: {lo} ~ {hi}")
        else:
            val = input_with_default("  请输入值", "")
            config["fixed_value"] = val
            config["mode"] = "fixed"

    else:
        # 自动生成
        config["mode"] = "auto"
        print(f"  ✅ {field_name} → 自动生成")

    return config


def interactive_configure_table(table_name, fields_def):
    """
    交互式配置一张表的字段

    Returns:
        dict: {field_name: {mode, fixed_value, ...}}
    """
    print_title(f"配置表: {table_name}")

    # 第一步：选择要生成哪些字段
    print("\n请选择要生成数据的字段（不需要的会留空 NULL）：\n")

    field_list = list(fields_def.keys())
    selected_fields = []

    for i, fname in enumerate(field_list, 1):
        fdef = fields_def[fname]
        label = fdef.get("label", fname)
        required = fdef.get("required", False)
        is_auto = fdef.get("auto", False)
        is_inherit = fdef.get("type") == "inherit"

        if required:
            print(f"  {i:2d}. ☑ {fname:40s} ({label}) — 必填")
            selected_fields.append(fname)
        elif is_inherit:
            print(f"  {i:2d}. ☑ {fname:40s} ({label}) — 继承自主表")
            selected_fields.append(fname)
        elif is_auto:
            print(f"  {i:2d}. ☑ {fname:40s} ({label}) — 自动计算")
            selected_fields.append(fname)
        else:
            default = "y" if i <= 10 else "n"  # 默认前10个字段选中
            val = input(f"  {i:2d}.   {fname:40s} ({label}) — 是否生成? (y/n) [{default}]: ").strip().lower()
            if val != "n" and (val == "y" or default == "y"):
                selected_fields.append(fname)

    if not selected_fields:
        print("\n⚠️  没有选择任何字段，将使用全部字段。")
        selected_fields = field_list

    print(f"\n已选择 {len(selected_fields)} 个字段: {', '.join(selected_fields)}")

    # 第二步：逐个配置字段
    print_title(f"详细配置: {table_name}")

    field_config = {}
    skip_auto = yes_no("\n自动计算的字段(费用、时间等)用默认方式生成？", "y")

    for fname in selected_fields:
        fdef = fields_def[fname]

        # 必填/继承/自动字段可以跳过详细配置
        if fdef.get("required") and fdef.get("type") != "choice":
            if yes_no(f"  {fname} ({fdef.get('label', '')}) 用默认自动生成?", "y"):
                field_config[fname] = {"mode": "auto"}
                continue

        if fdef.get("auto") and skip_auto:
            field_config[fname] = {"mode": "auto"}
            continue

        if fdef.get("type") == "inherit":
            field_config[fname] = {"mode": "auto"}
            continue

        config = configure_field_interactive(fname, fdef)
        field_config[fname] = config

    # 未选中的字段设为 null
    for fname in field_list:
        if fname not in field_config:
            fdef = fields_def[fname]
            if fdef.get("required") or fdef.get("auto") or fdef.get("type") == "inherit":
                field_config[fname] = {"mode": "auto"}
            else:
                field_config[fname] = {"mode": "null"}

    return field_config


def interactive_main():
    """交互式主流程"""
    print_separator()
    print("  🚀 亚马逊 FBA 入库货件 —— 交互式数据生成器")
    print_separator()

    # 1. 基本参数
    print("\n【第一步：基本参数】\n")
    count = int(input_with_default("  生成货件数量", "50"))
    min_items = int(input_with_default("  每个货件最少 SKU 数", "1"))
    max_items = int(input_with_default("  每个货件最多 SKU 数", "5"))

    # 2. 是否使用已保存的模板
    template = None
    template_from_excel_dir = False
    print("\n  加载配置模板的方式:")
    print("    1. 不使用模板（手动交互配置）")
    print("    2. 加载 JSON 模板")
    print("    3. 加载 Excel 模板")
    template_choice = input_with_default("  请选择", "1")
    if template_choice == "2":
        template_name = input_with_default("  模板名称", "default")
        template = load_template(template_name)
        if template:
            print(f"  ✅ 已加载模板: {template_name}")
        else:
            print(f"  ⚠️  模板不存在，将进入手动配置")
    elif template_choice == "3":
        template = _interactive_pick_excel_template()
        template_from_excel_dir = template is not None

    # 3. 配置货件主表字段
    if template and "shipment" in template:
        shipment_config = template["shipment"]
        print(f"\n  使用模板中的货件主表配置")
    else:
        shipment_config = interactive_configure_table(
            "mws_fi_data_inbound_shipment", SHIPMENT_FIELDS
        )

    # 4. 配置明细表字段
    if template and "item" in template:
        item_config = template["item"]
        print(f"\n  使用模板中的明细表配置")
    else:
        item_config = interactive_configure_table(
            "mws_fi_data_inbound_shipment_item", ITEM_FIELDS
        )

    # 5. 保存模板（如果配置来自 excel_templates 目录，跳过保存）
    if template_from_excel_dir:
        print("\n  ✅ 配置来自 excel_templates 目录，无需重复保存")
    elif yes_no("\n是否保存当前配置为模板？（下次可复用）", "y"):
        print("  保存格式:")
        print("    1. JSON 模板")
        print("    2. Excel 模板")
        print("    3. 两种都保存")
        save_choice = input_with_default("  请选择", "2")
        template_name = input_with_default("  模板名称", "default")
        if save_choice in ("1", "3"):
            save_template(template_name, shipment_config, item_config)
        if save_choice in ("2", "3"):
            excel_path = template_name if template_name.endswith('.xlsx') else f"{template_name}.xlsx"
            save_template_excel(os.path.join(EXCEL_TEMPLATE_DIR, excel_path), shipment_config, item_config)

    # 6. 输出方式
    print("\n【输出方式】\n")
    print("  1. CSV 文件")
    print("  2. SQL 文件")
    print("  3. 直连 MySQL 写入")
    output_choice = input_with_default("  请选择", "1")

    db_create_table = False
    db_truncate = False
    db_dry_run = False

    if output_choice == "1":
        output_mode = "csv"
        output_dir = input_with_default("  输出目录", "./data")
    elif output_choice == "2":
        output_mode = "sql"
        output_dir = input_with_default("  SQL 文件路径", "./data/insert_data.sql")
    else:
        output_mode = "db"
        # 尝试加载已保存的数据库配置
        saved_cfg = load_db_config()
        default_host = saved_cfg["host"]
        default_port = str(saved_cfg["port"])
        default_user = saved_cfg["user"]
        default_db = saved_cfg["database"]

        db_host = input_with_default("  MySQL 主机", default_host)
        db_port = int(input_with_default("  MySQL 端口", default_port))
        db_user = input_with_default("  MySQL 用户", default_user)
        db_pass = input("  MySQL 密码: ").strip()
        db_name = input_with_default("  数据库名", default_db)

        # 数据库操作选项
        if yes_no("  是否自动创建表（IF NOT EXISTS）?", "y"):
            db_create_table = True
        if yes_no("  是否清空已有数据?", "n"):
            db_truncate = True

        # 保存数据库配置
        if not os.path.exists(DB_CONFIG_FILE):
            if yes_no("  是否保存数据库配置（下次免输入）?", "y"):
                save_db_config({"host": db_host, "port": db_port, "user": db_user, "password": db_pass, "database": db_name})

    # 7. 开始生成
    print_title("开始生成数据")

    generator = InboundDataGenerator()
    shipments, items = generator.generate_batch(
        count=count,
        shipment_config=shipment_config,
        item_config=item_config,
        min_items=min_items,
        max_items=max_items,
    )

    # 统计
    print(f"\n📊 数据统计:")
    print(f"  货件总数: {len(shipments)}")
    print(f"  明细总数: {len(items)}")
    if shipments:
        print(f"  平均每货件 SKU: {len(items) / len(shipments):.1f}")

    # 输出
    if output_mode == "csv":
        export_to_csv(shipments, items, output_dir)
    elif output_mode == "sql":
        export_to_sql(shipments, items, output_dir)
    elif output_mode == "db":
        insert_to_mysql(shipments, items, db_host, db_port, db_user, db_pass, db_name,
                        truncate=db_truncate, create_table=db_create_table, dry_run=db_dry_run)

    print("\n🎉 完成！")


def interactive_main_with_template(args, template):
    """使用模板的半交互式模式：模板提供配置，只需输入数量和输出方式"""
    print_separator()
    print("  🚀 亚马逊 FBA 入库货件 —— 模板模式")
    print_separator()

    shipment_config = template.get("shipment", {})
    item_config = template.get("item", {})

    # 只需输入数量
    print("\n【基本参数】\n")
    count = int(input_with_default("  生成货件数量", str(args.count)))
    min_items = int(input_with_default("  每个货件最少 SKU 数", str(args.min_items)))
    max_items = int(input_with_default("  每个货件最多 SKU 数", str(args.max_items)))

    # 输出方式
    print("\n【输出方式】\n")
    print("  1. CSV 文件")
    print("  2. SQL 文件")
    print("  3. 直连 MySQL 写入")
    output_choice = input_with_default("  请选择", "1")

    db_create_table = False
    db_truncate = False
    db_dry_run = False

    if output_choice == "1":
        output_mode = "csv"
        output_dir = input_with_default("  输出目录", "./data")
    elif output_choice == "2":
        output_mode = "sql"
        output_dir = input_with_default("  SQL 文件路径", "./data/insert_data.sql")
    else:
        output_mode = "db"
        # 尝试加载已保存的数据库配置
        saved_cfg = load_db_config()
        default_host = saved_cfg["host"]
        default_port = str(saved_cfg["port"])
        default_user = saved_cfg["user"]
        default_db = saved_cfg["database"]

        db_host = input_with_default("  MySQL 主机", default_host)
        db_port = int(input_with_default("  MySQL 端口", default_port))
        db_user = input_with_default("  MySQL 用户", default_user)
        db_pass = input("  MySQL 密码: ").strip()
        db_name = input_with_default("  数据库名", default_db)

        # 数据库操作选项
        if yes_no("  是否自动创建表（IF NOT EXISTS）?", "y"):
            db_create_table = True
        if yes_no("  是否清空已有数据?", "n"):
            db_truncate = True

    # 生成
    print_title("开始生成数据")

    generator = InboundDataGenerator()
    shipments, items = generator.generate_batch(
        count=count,
        shipment_config=shipment_config,
        item_config=item_config,
        min_items=min_items,
        max_items=max_items,
    )

    print(f"\n📊 数据统计:")
    print(f"  货件总数: {len(shipments)}")
    print(f"  明细总数: {len(items)}")
    if shipments:
        print(f"  平均每货件 SKU: {len(items) / len(shipments):.1f}")

    if output_mode == "csv":
        export_to_csv(shipments, items, output_dir)
    elif output_mode == "sql":
        export_to_sql(shipments, items, output_dir)
    elif output_mode == "db":
        insert_to_mysql(shipments, items, db_host, db_port, db_user, db_pass, db_name,
                        truncate=db_truncate, create_table=db_create_table, dry_run=db_dry_run)

    print("\n🎉 完成！")


# ============================================================
# 模板管理
# ============================================================

def save_template(name, shipment_config, item_config):
    """保存配置模板为 JSON"""
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    filepath = os.path.join(TEMPLATE_DIR, f"{name}.json")
    data = {
        "name": name,
        "created_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "shipment": shipment_config,
        "item": item_config,
    }
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON 模板已保存: {filepath}")


def load_template(name):
    """加载 JSON 配置模板"""
    filepath = os.path.join(TEMPLATE_DIR, f"{name}.json")
    if not os.path.exists(filepath):
        return None
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def list_templates():
    """列出所有模板（JSON + Excel）"""
    templates = []

    # JSON 模板
    if os.path.exists(TEMPLATE_DIR):
        for f in os.listdir(TEMPLATE_DIR):
            if f.endswith('.json'):
                filepath = os.path.join(TEMPLATE_DIR, f)
                with open(filepath, 'r', encoding='utf-8') as fp:
                    data = json.load(fp)
                templates.append({
                    "name": data.get("name", f[:-5]),
                    "created_at": data.get("created_at", "未知"),
                    "filepath": filepath,
                    "format": "json",
                })

    # Excel 模板
    templates.extend(list_excel_templates())

    return templates


# ============================================================
# Excel 模板管理
# ============================================================

def _resolve_excel_template_path(name):
    """
    解析 Excel 模板路径：
    - 如果是完整路径且文件存在，直接返回
    - 否则自动到 excel_templates/ 目录下查找
    """
    # 已经是完整路径
    if os.path.exists(name):
        return name

    # 在 excel_templates 目录下查找
    candidates = [
        os.path.join(EXCEL_TEMPLATE_DIR, name),
        os.path.join(EXCEL_TEMPLATE_DIR, name + '.xlsx'),
        os.path.join(EXCEL_TEMPLATE_DIR, name + '.xls'),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path

    # 都找不到，返回原始路径（让 load_template_excel 报错）
    return name


def _get_field_choices_str(fdef):
    """获取字段可选值的描述字符串"""
    choices = fdef.get("choices", [])
    if isinstance(choices, dict):
        return ",".join(f"{k}({v})" for k, v in choices.items())
    elif isinstance(choices, list):
        return ",".join(str(c) for c in choices)
    return ""


def list_excel_templates():
    """列出 excel_templates 目录下的所有 Excel 模板"""
    os.makedirs(EXCEL_TEMPLATE_DIR, exist_ok=True)
    templates = []
    for f in sorted(os.listdir(EXCEL_TEMPLATE_DIR)):
        if f.endswith(('.xlsx', '.xls')):
            filepath = os.path.join(EXCEL_TEMPLATE_DIR, f)
            mtime = os.path.getmtime(filepath)
            created_at = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
            templates.append({
                "name": f,
                "filepath": filepath,
                "created_at": created_at,
                "format": "excel",
            })
    return templates


def _interactive_pick_excel_template():
    """交互式选择 excel_templates 目录下的 Excel 模板"""
    templates = list_excel_templates()
    if not templates:
        print("\n  ⚠️  excel_templates 目录下没有 Excel 模板文件")
        if yes_no("  是否现在导出一个空白模板？", "y"):
            os.makedirs(EXCEL_TEMPLATE_DIR, exist_ok=True)
            default_path = os.path.join(EXCEL_TEMPLATE_DIR, "template_config.xlsx")
            export_template_excel(default_path)
            print(f"  请编辑 {default_path} 后重新运行")
        return None

    print("\n  excel_templates 目录下的 Excel 模板：")
    for i, t in enumerate(templates, 1):
        print(f"    {i}. {t['name']}  (修改于 {t['created_at']})")

    choice = input_with_default("  请选择模板编号", "1")
    try:
        idx = int(choice) - 1
        if 0 <= idx < len(templates):
            selected = templates[idx]
            template = load_template_excel(selected['filepath'])
            return template
        else:
            print(f"  ❌ 无效选择")
            return None
    except ValueError:
        print(f"  ❌ 无效输入")
        return None


def export_template_excel(output_path=None):
    """
    导出空白 Excel 配置模板，用户可在 Excel 中编辑后使用

    Excel 结构：
      Sheet1 "货件主表" — 字段配置
      Sheet2 "货件明细" — 字段配置
      Sheet3 "说明" — 使用说明
    """
    if not HAS_OPENPYXL:
        print("缺少 openpyxl 库，请执行: pip install openpyxl")
        sys.exit(1)

    # 默认保存到 excel_templates 目录
    if output_path is None:
        os.makedirs(EXCEL_TEMPLATE_DIR, exist_ok=True)
        output_path = os.path.join(EXCEL_TEMPLATE_DIR, "template_config.xlsx")

    wb = openpyxl.Workbook()

    # ---- Sheet1: 货件主表 ----
    ws1 = wb.active
    ws1.title = "货件主表"

    # 表头
    headers = ["字段名", "中文名", "是否生成", "生成模式", "固定值", "范围最小值", "范围最大值", "可选项(逗号分隔)"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    # 填充字段
    for i, (fname, fdef) in enumerate(SHIPMENT_FIELDS.items(), 2):
        is_required = fdef.get("required", False)
        is_auto = fdef.get("auto", False)
        is_inherit = fdef.get("type") == "inherit"
        is_nullable = fdef.get("nullable", False)

        # 字段名
        ws1.cell(row=i, column=1, value=fname)
        # 中文名
        ws1.cell(row=i, column=2, value=fdef.get("label", ""))
        # 是否生成
        if is_required or is_auto or is_inherit:
            ws1.cell(row=i, column=3, value="是")
            cell = ws1.cell(row=i, column=3)
            cell.fill = openpyxl.styles.PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        else:
            ws1.cell(row=i, column=3, value="是")  # 默认是
        # 生成模式
        if is_auto or is_inherit:
            ws1.cell(row=i, column=4, value="auto")
        else:
            ws1.cell(row=i, column=4, value="auto")
        # 固定值
        ws1.cell(row=i, column=5, value="")
        # 范围最小值
        rng = fdef.get("range", [])
        ws1.cell(row=i, column=6, value=rng[0] if rng else "")
        # 范围最大值
        ws1.cell(row=i, column=7, value=rng[1] if rng else "")
        # 可选项
        choices_str = _get_field_choices_str(fdef)
        ws1.cell(row=i, column=8, value=choices_str)

        # 必填行标灰
        if is_required:
            for col in range(1, 9):
                cell = ws1.cell(row=i, column=col)
                if not cell.fill or cell.fill.start_color.rgb == "00000000":
                    cell.fill = openpyxl.styles.PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # 列宽
    col_widths = [38, 18, 10, 12, 25, 14, 14, 50]
    for idx, w in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    # ---- Sheet2: 货件明细 ----
    ws2 = wb.create_sheet("货件明细")

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    for i, (fname, fdef) in enumerate(ITEM_FIELDS.items(), 2):
        is_required = fdef.get("required", False)
        is_auto = fdef.get("auto", False)
        is_inherit = fdef.get("type") == "inherit"

        ws2.cell(row=i, column=1, value=fname)
        ws2.cell(row=i, column=2, value=fdef.get("label", ""))
        if is_required or is_auto or is_inherit:
            ws2.cell(row=i, column=3, value="是")
        else:
            ws2.cell(row=i, column=3, value="是")
        if is_auto or is_inherit:
            ws2.cell(row=i, column=4, value="auto")
        else:
            ws2.cell(row=i, column=4, value="auto")
        ws2.cell(row=i, column=5, value="")
        rng = fdef.get("range", [])
        ws2.cell(row=i, column=6, value=rng[0] if rng else "")
        ws2.cell(row=i, column=7, value=rng[1] if rng else "")
        choices_str = _get_field_choices_str(fdef)
        ws2.cell(row=i, column=8, value=choices_str)

        if is_required or is_inherit:
            for col in range(1, 9):
                cell = ws2.cell(row=i, column=col)
                if not cell.fill or cell.fill.start_color.rgb == "00000000":
                    cell.fill = openpyxl.styles.PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for idx, w in enumerate(col_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    # ---- Sheet3: 说明 ----
    ws3 = wb.create_sheet("说明")
    instructions = [
        ["Excel 配置模板使用说明", ""],
        ["", ""],
        ["列说明:", ""],
        ["字段名", "数据库字段名，请勿修改"],
        ["中文名", "字段中文描述，仅供参考"],
        ["是否生成", "填 是 或 否。否 = 该字段留空(NULL)"],
        ["生成模式", "auto=自动随机生成, fixed=使用固定值, range=范围内随机, null=留空"],
        ["固定值", "生成模式=fixed 时，填写每条数据都使用的固定值"],
        ["范围最小值", "生成模式=range 时，填写随机范围的最小值"],
        ["范围最大值", "生成模式=range 时，填写随机范围的最大值"],
        ["可选项", "生成模式=range 且字段类型=choice/string 时，填写可选值(逗号分隔)，将从中随机选"],
        ["", ""],
        ["示例:", ""],
        ["只想造某个卖家的数据", "SELLER_ID 行: 生成模式=fixed, 固定值=AXXXXXXXXXXXXX"],
        ["只想造已完成状态的数据", "SHIPMENT_STATUS 行: 生成模式=fixed, 固定值=7"],
        ["发货数量控制在100~300之间", "QUANTITY_SHIPPED 行: 生成模式=range, 范围最小值=100, 范围最大值=300"],
        ["只要美国市场的数据", "MARKETPLACE_ID 行: 生成模式=fixed, 固定值=ATVPDKIKX0DER"],
        ["不要地址信息", "地址类字段: 是否生成=否"],
        ["", ""],
        ["使用方式:", ""],
        ["1. 编辑此 Excel 文件中的配置"],
        ["2. 保存后执行: python generate_inbound_data.py --template-excel excel_templates/template_config.xlsx --count 50 --output csv"],
    ]
    for i, row in enumerate(instructions, 1):
        for j, val in enumerate(row, 1):
            ws3.cell(row=i, column=j, value=val)
        if i == 1 or i == 3 or i == 13 or i == 19:
            ws3.cell(row=i, column=1).font = openpyxl.styles.Font(bold=True, size=12)
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 70

    wb.save(output_path)
    print(f"✅ Excel 模板已导出: {output_path}")
    print(f"   请用 Excel 打开编辑，然后使用 --template-excel 参数加载")


def load_template_excel(filepath):
    """
    从 Excel 文件加载配置模板

    Returns:
        dict: {"shipment": {...}, "item": {...}}
    """
    if not HAS_OPENPYXL:
        print("缺少 openpyxl 库，请执行: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(filepath):
        print(f"❌ 文件不存在: {filepath}")
        return None

    wb = openpyxl.load_workbook(filepath)

    def parse_sheet(ws, fields_def):
        """解析一个 Sheet 的配置"""
        config = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            fname = str(row[0]).strip()
            if fname not in fields_def:
                continue

            is_generate = str(row[2]).strip() if row[2] else "是"
            mode = str(row[3]).strip() if row[3] else "auto"
            fixed_value = str(row[4]).strip() if row[4] else ""
            range_min = row[5]
            range_max = row[6]
            choices_str = str(row[7]).strip() if row[7] else ""

            if is_generate in ("否", "n", "N", "no", "NO"):
                fdef = fields_def[fname]
                if fdef.get("required") or fdef.get("auto") or fdef.get("type") == "inherit":
                    config[fname] = {"mode": "auto"}
                else:
                    config[fname] = {"mode": "null"}
                continue

            if mode == "fixed" and fixed_value:
                config[fname] = {"mode": "fixed", "fixed_value": fixed_value}
            elif mode == "range":
                cfg = {"mode": "range"}
                if range_min is not None:
                    cfg["range_min"] = range_min
                if range_max is not None:
                    cfg["range_max"] = range_max
                if choices_str:
                    cfg["choices"] = [c.strip() for c in choices_str.split(",") if c.strip()]
                config[fname] = cfg
            elif mode == "null":
                config[fname] = {"mode": "null"}
            else:
                config[fname] = {"mode": "auto"}

        # 填充缺失字段
        for fname, fdef in fields_def.items():
            if fname not in config:
                if fdef.get("required") or fdef.get("auto") or fdef.get("type") == "inherit":
                    config[fname] = {"mode": "auto"}
                else:
                    config[fname] = {"mode": "null"}

        return config

    shipment_config = parse_sheet(wb["货件主表"], SHIPMENT_FIELDS)
    item_config = parse_sheet(wb["货件明细"], ITEM_FIELDS)

    print(f"✅ 已加载 Excel 模板: {filepath}")
    return {"shipment": shipment_config, "item": item_config}


def save_template_excel(filepath, shipment_config, item_config):
    """将当前配置保存为 Excel 模板"""
    if not HAS_OPENPYXL:
        print("缺少 openpyxl 库，请执行: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    def write_sheet(ws, title, fields_def, config_data, header_color):
        ws.title = title
        headers = ["字段名", "中文名", "是否生成", "生成模式", "固定值", "范围最小值", "范围最大值", "可选项(逗号分隔)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        for i, (fname, fdef) in enumerate(fields_def.items(), 2):
            cfg = config_data.get(fname, {"mode": "auto"})
            mode = cfg.get("mode", "auto")

            ws.cell(row=i, column=1, value=fname)
            ws.cell(row=i, column=2, value=fdef.get("label", ""))
            ws.cell(row=i, column=3, value="否" if mode == "null" else "是")
            ws.cell(row=i, column=4, value=mode)
            ws.cell(row=i, column=5, value=cfg.get("fixed_value", ""))
            ws.cell(row=i, column=6, value=cfg.get("range_min", ""))
            ws.cell(row=i, column=7, value=cfg.get("range_max", ""))
            ws.cell(row=i, column=8, value=",".join(cfg.get("choices", [])))

        col_widths = [38, 18, 10, 12, 25, 14, 14, 50]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    write_sheet(wb.active, "货件主表", SHIPMENT_FIELDS, shipment_config, "4472C4")
    write_sheet(wb.create_sheet("货件明细"), "货件明细", ITEM_FIELDS, item_config, "548235")

    wb.save(filepath)
    print(f"  ✅ Excel 模板已保存: {filepath}")


# ============================================================
# 输出方式
# ============================================================

def export_to_csv(shipments, items, output_dir='.'):
    os.makedirs(output_dir, exist_ok=True)

    # 过滤掉全为 None 的列
    def filter_columns(data):
        if not data:
            return data
        all_keys = list(data[0].keys())
        keep_keys = []
        for k in all_keys:
            has_value = any(row.get(k) is not None for row in data)
            if has_value:
                keep_keys.append(k)
        return [{k: row.get(k) for k in keep_keys} for row in data]

    shipments = filter_columns(shipments)
    items = filter_columns(items)

    shipment_file = os.path.join(output_dir, 'mws_fi_data_inbound_shipment.csv')
    if shipments:
        keys = shipments[0].keys()
        with open(shipment_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(shipments)
        print(f"✅ 货件主表已导出: {shipment_file} ({len(shipments)} 条)")

    item_file = os.path.join(output_dir, 'mws_fi_data_inbound_shipment_item.csv')
    if items:
        keys = items[0].keys()
        with open(item_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(items)
        print(f"✅ 货件明细已导出: {item_file} ({len(items)} 条)")


def export_to_sql(shipments, items, sql_file='insert_data.sql'):
    dir_path = os.path.dirname(sql_file)
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)

    with open(sql_file, 'w', encoding='utf-8') as f:
        f.write("-- 亚马逊 FBA 入库货件测试数据\n")
        f.write(f"-- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"-- 货件: {len(shipments)} 条, 明细: {len(items)} 条\n\n")

        if shipments:
            f.write("-- 货件主表\n")
            for s in shipments:
                # 只插入非 None 的字段
                cols_vals = [(k, v) for k, v in s.items() if v is not None]
                if cols_vals:
                    cols = ', '.join(k for k, _ in cols_vals)
                    vals = ', '.join(_sql_value(v) for _, v in cols_vals)
                    f.write(f"INSERT INTO `mws_fi_data_inbound_shipment` ({cols}) VALUES ({vals});\n")

        if items:
            f.write("\n-- 货件明细\n")
            for item in items:
                cols_vals = [(k, v) for k, v in item.items() if v is not None]
                if cols_vals:
                    cols = ', '.join(k for k, _ in cols_vals)
                    vals = ', '.join(_sql_value(v) for _, v in cols_vals)
                    f.write(f"INSERT INTO `mws_fi_data_inbound_shipment_item` ({cols}) VALUES ({vals});\n")

    print(f"✅ SQL 文件已导出: {sql_file}")


def _sql_value(value):
    if value is None:
        return 'NULL'
    elif isinstance(value, bool):
        return '1' if value else '0'
    elif isinstance(value, (int, float)):
        return str(value)
    else:
        escaped = str(value).replace("'", "\\'").replace("\\", "\\\\")
        return f"'{escaped}'"


# ============================================================
# 数据库配置管理
# ============================================================

# 数据库配置文件路径（与脚本同目录）
DB_CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "db_config.json")

# MySQL 字段类型映射：根据字段定义自动推导 SQL 列类型
_MYSQL_TYPE_MAP = {
    "string": "VARCHAR(255)",
    "choice": "VARCHAR(50)",
    "int": "INT",
    "float": "DECIMAL(12,2)",
    "datetime": "DATETIME",
    "inherit": "VARCHAR(255)",  # 继承字段类型由源字段决定，此处用通用类型
}


def load_db_config(config_file=None):
    """
    从 db_config.json 加载数据库连接配置。
    配置文件格式：
    {
        "host": "127.0.0.1",
        "port": 3306,
        "user": "root",
        "password": "your_password",
        "database": "test"
    }

    也支持从环境变量读取（优先级更高）：
        DB_HOST, DB_PORT, DB_USER, DB_PASSWORD, DB_DATABASE
    """
    config = {
        "host": os.environ.get("DB_HOST", "127.0.0.1"),
        "port": int(os.environ.get("DB_PORT", "3306")),
        "user": os.environ.get("DB_USER", "root"),
        "password": os.environ.get("DB_PASSWORD", ""),
        "database": os.environ.get("DB_DATABASE", "test"),
    }

    # 从配置文件读取（环境变量未设置时作为默认值）
    cfg_path = config_file or DB_CONFIG_FILE
    if os.path.exists(cfg_path):
        try:
            with open(cfg_path, 'r', encoding='utf-8') as f:
                file_config = json.load(f)
            # 环境变量优先，配置文件补缺
            for key in ("host", "port", "user", "password", "database"):
                if key in file_config:
                    # 环境变量未设置时才用文件中的值
                    env_key = f"DB_{key.upper()}"
                    if env_key not in os.environ:
                        config[key] = file_config[key]
            print(f"  📂 已加载数据库配置: {cfg_path}")
        except Exception as e:
            print(f"  ⚠️  读取数据库配置文件失败: {e}，使用默认值")
    else:
        if config_file:
            print(f"  ⚠️  指定的配置文件不存在: {cfg_path}")

    # 确保 port 是整数
    config["port"] = int(config["port"])
    return config


def save_db_config(config, config_file=None):
    """保存数据库连接配置到 db_config.json"""
    cfg_path = config_file or DB_CONFIG_FILE
    try:
        with open(cfg_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"  ✅ 数据库配置已保存: {cfg_path}")
        print(f"     ⚠️  配置文件包含密码，请勿提交到版本控制！建议在 .gitignore 中添加 db_config.json")
    except Exception as e:
        print(f"  ❌ 保存配置文件失败: {e}")


def test_mysql_connection(host, port, user, password, database):
    """测试 MySQL 数据库连接是否可用"""
    try:
        import pymysql
    except ImportError:
        print("❌ 缺少 pymysql，请执行: pip install pymysql")
        return False

    try:
        conn = pymysql.connect(
            host=host, port=port, user=user,
            password=password, database=database,
            charset='utf8mb4', connect_timeout=10,
        )
        cursor = conn.cursor()
        cursor.execute("SELECT VERSION()")
        version = cursor.fetchone()[0]
        cursor.close()
        conn.close()
        print(f"  ✅ 数据库连接成功！MySQL 版本: {version}")
        return True
    except Exception as e:
        print(f"  ❌ 数据库连接失败: {e}")
        return False


def generate_create_table_ddl(table_name, fields_def, if_not_exists=True):
    """
    根据字段定义自动生成 CREATE TABLE 语句。

    Args:
        table_name: 表名
        fields_def: 字段定义字典 (SHIPMENT_FIELDS / ITEM_FIELDS)
        if_not_exists: 是否添加 IF NOT EXISTS

    Returns:
        str: CREATE TABLE DDL 语句
    """
    exists_clause = "IF NOT EXISTS " if if_not_exists else ""
    lines = []

    for fname, fdef in fields_def.items():
        ftype = fdef.get("type", "string")
        mysql_type = _MYSQL_TYPE_MAP.get(ftype, "VARCHAR(255)")

        # 继承字段根据源字段类型推导
        if ftype == "inherit":
            inherit_from = fdef.get("inherit_from", "shipment")
            if inherit_from == "shipment" and fname in SHIPMENT_FIELDS:
                source_type = SHIPMENT_FIELDS[fname].get("type", "string")
                mysql_type = _MYSQL_TYPE_MAP.get(source_type, "VARCHAR(255)")

        nullable = fdef.get("nullable", False)
        required = fdef.get("required", False)

        if required:
            null_clause = " NOT NULL"
        elif nullable:
            null_clause = ""
        else:
            null_clause = ""

        default_clause = ""
        if nullable and not required:
            default_clause = " DEFAULT NULL"

        lines.append(f"  `{fname}` {mysql_type}{null_clause}{default_clause}")

    # 主键：使用 SHIPMENT_ID 作为货件主表主键
    if table_name == "mws_fi_data_inbound_shipment":
        lines.append(f"  PRIMARY KEY (`SHIPMENT_ID`)")
    elif table_name == "mws_fi_data_inbound_shipment_item":
        # 明细表使用联合主键
        lines.append(f"  PRIMARY KEY (`SHIPMENT_ID`, `SELLER_SKU`)")

    columns_sql = ",\n".join(lines)
    ddl = f"CREATE TABLE {exists_clause}`{table_name}` (\n{columns_sql}\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;"
    return ddl


def create_tables_if_needed(host, port, user, password, database):
    """自动创建目标表（如果不存在）"""
    try:
        import pymysql
    except ImportError:
        print("❌ 缺少 pymysql，请执行: pip install pymysql")
        return False

    try:
        conn = pymysql.connect(
            host=host, port=port, user=user,
            password=password, database=database, charset='utf8mb4',
        )
        cursor = conn.cursor()

        # 货件主表
        ddl1 = generate_create_table_ddl("mws_fi_data_inbound_shipment", SHIPMENT_FIELDS)
        print(f"\n📦 创建货件主表...")
        print(f"  执行 DDL:\n{ddl1}\n")
        cursor.execute(ddl1)
        print(f"  ✅ 货件主表已就绪")

        # 明细表
        ddl2 = generate_create_table_ddl("mws_fi_data_inbound_shipment_item", ITEM_FIELDS)
        print(f"\n📋 创建明细表...")
        print(f"  执行 DDL:\n{ddl2}\n")
        cursor.execute(ddl2)
        print(f"  ✅ 明细表已就绪")

        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Exception as e:
        print(f"  ❌ 创建表失败: {e}")
        return False


def truncate_tables(host, port, user, password, database):
    """清空目标表数据（保留表结构）"""
    try:
        import pymysql
    except ImportError:
        print("❌ 缺少 pymysql，请执行: pip install pymysql")
        return False

    try:
        conn = pymysql.connect(
            host=host, port=port, user=user,
            password=password, database=database, charset='utf8mb4',
        )
        cursor = conn.cursor()

        # 先删除明细表（外键依赖），再删除主表
        print(f"\n🗑️  清空明细表...")
        cursor.execute("TRUNCATE TABLE `mws_fi_data_inbound_shipment_item`")
        print(f"  ✅ 明细表已清空")

        print(f"\n🗑️  清空货件主表...")
        cursor.execute("TRUNCATE TABLE `mws_fi_data_inbound_shipment`")
        print(f"  ✅ 货件主表已清空")

        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Exception as e:
        print(f"  ❌ 清空表失败: {e}")
        return False


def insert_to_mysql(shipments, items, host, port, user, password, database,
                    batch_size=500, truncate=False, create_table=False,
                    dry_run=False, retry_count=3, retry_delay=2):
    """
    将生成的数据写入 MySQL 数据库。

    Args:
        shipments: 货件主表数据列表
        items: 明细表数据列表
        host/port/user/password/database: 数据库连接参数
        batch_size: 每批提交的行数
        truncate: 是否在插入前清空表
        create_table: 是否自动建表
        dry_run: 仅打印 SQL 不实际执行
        retry_count: 连接失败重试次数
        retry_delay: 重试间隔（秒）
    """
    try:
        import pymysql
    except ImportError:
        print("❌ 缺少 pymysql，请执行: pip install pymysql")
        sys.exit(1)

    # 连接测试（带重试）
    conn = None
    last_error = None
    for attempt in range(1, retry_count + 1):
        try:
            conn = pymysql.connect(
                host=host, port=port, user=user,
                password=password, database=database,
                charset='utf8mb4', connect_timeout=10,
            )
            break
        except Exception as e:
            last_error = e
            if attempt < retry_count:
                print(f"  ⚠️  连接失败（第 {attempt}/{retry_count} 次），{retry_delay}秒后重试...")
                time.sleep(retry_delay)
            else:
                print(f"  ❌ 连接失败，已重试 {retry_count} 次: {last_error}")
                sys.exit(1)

    try:
        cursor = conn.cursor()

        # 测试连接
        cursor.execute("SELECT VERSION()")
        version = cursor.fetchone()[0]
        print(f"\n🔌 已连接 MySQL {version} @ {host}:{port}/{database}")

        # 自动建表
        if create_table:
            ddl1 = generate_create_table_ddl("mws_fi_data_inbound_shipment", SHIPMENT_FIELDS)
            ddl2 = generate_create_table_ddl("mws_fi_data_inbound_shipment_item", ITEM_FIELDS)
            if dry_run:
                print(f"\n[DRY-RUN] 将执行以下建表语句:")
                print(f"\n--- 货件主表 DDL ---\n{ddl1}\n")
                print(f"--- 明细表 DDL ---\n{ddl2}\n")
            else:
                print(f"\n📦 自动建表...")
                cursor.execute(ddl1)
                print(f"  ✅ 货件主表已就绪")
                cursor.execute(ddl2)
                print(f"  ✅ 明细表已就绪")
                conn.commit()

        # 清空表
        if truncate:
            if dry_run:
                print(f"\n[DRY-RUN] 将执行以下清空操作:")
                print(f"  TRUNCATE TABLE `mws_fi_data_inbound_shipment_item`")
                print(f"  TRUNCATE TABLE `mws_fi_data_inbound_shipment`")
            else:
                print(f"\n🗑️  清空目标表...")
                cursor.execute("TRUNCATE TABLE `mws_fi_data_inbound_shipment_item`")
                cursor.execute("TRUNCATE TABLE `mws_fi_data_inbound_shipment`")
                conn.commit()
                print(f"  ✅ 表已清空")

        # dry-run 模式：打印前几条 INSERT 语句
        if dry_run:
            print(f"\n[DRY-RUN] 将执行以下 INSERT 语句（仅展示前3条）:")
            if shipments:
                print(f"\n--- 货件主表（共 {len(shipments)} 条）---")
                for s in shipments[:3]:
                    cols_vals = [(k, v) for k, v in s.items() if v is not None]
                    if cols_vals:
                        cols = ', '.join(k for k, _ in cols_vals)
                        vals = ', '.join(_sql_value(v) for _, v in cols_vals)
                        print(f"  INSERT INTO `mws_fi_data_inbound_shipment` ({cols}) VALUES ({vals});")
                if len(shipments) > 3:
                    print(f"  ... 省略 {len(shipments) - 3} 条")
            if items:
                print(f"\n--- 明细表（共 {len(items)} 条）---")
                for item in items[:3]:
                    cols_vals = [(k, v) for k, v in item.items() if v is not None]
                    if cols_vals:
                        cols = ', '.join(k for k, _ in cols_vals)
                        vals = ', '.join(_sql_value(v) for _, v in cols_vals)
                        print(f"  INSERT INTO `mws_fi_data_inbound_shipment_item` ({cols}) VALUES ({vals});")
                if len(items) > 3:
                    print(f"  ... 省略 {len(items) - 3} 条")
            print(f"\n[DRY-RUN] 模式结束，未实际写入数据。")
            return

        # 动态构建 INSERT
        if shipments:
            print(f"\n📦 插入货件主表...")
            for i, s in enumerate(shipments):
                cols_vals = [(k, v) for k, v in s.items() if v is not None]
                if cols_vals:
                    cols = ', '.join(k for k, _ in cols_vals)
                    placeholders = ', '.join(['%s'] * len(cols_vals))
                    vals = [v for _, v in cols_vals]
                    sql = f"INSERT INTO `mws_fi_data_inbound_shipment` ({cols}) VALUES ({placeholders})"
                    cursor.execute(sql, vals)
                if (i + 1) % batch_size == 0:
                    conn.commit()
                    print(f"  已插入 {i + 1}/{len(shipments)} 条")
            conn.commit()
            print(f"  ✅ 货件主表完成: {len(shipments)} 条")

        if items:
            print(f"\n📋 插入货件明细...")
            for i, item in enumerate(items):
                cols_vals = [(k, v) for k, v in item.items() if v is not None]
                if cols_vals:
                    cols = ', '.join(k for k, _ in cols_vals)
                    placeholders = ', '.join(['%s'] * len(cols_vals))
                    vals = [v for _, v in cols_vals]
                    sql = f"INSERT INTO `mws_fi_data_inbound_shipment_item` ({cols}) VALUES ({placeholders})"
                    cursor.execute(sql, vals)
                if (i + 1) % batch_size == 0:
                    conn.commit()
                    print(f"  已插入 {i + 1}/{len(items)} 条")
            conn.commit()
            print(f"  ✅ 明细表完成: {len(items)} 条")

        print(f"\n✅ 数据写入完成！")

    except Exception as e:
        if conn:
            conn.rollback()
        print(f"❌ 写入失败: {e}")
        raise
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# ============================================================
# 命令行模式（非交互式，向后兼容）
# ============================================================

def cli_main(args):
    """命令行模式：快速生成，全部字段自动生成"""
    shipment_config = {fname: {"mode": "auto"} for fname in SHIPMENT_FIELDS}
    item_config = {fname: {"mode": "auto"} for fname in ITEM_FIELDS}
    cli_main_with_config(args, shipment_config, item_config)


def cli_main_with_config(args, shipment_config, item_config):
    """命令行模式：使用指定配置生成"""
    print_separator()
    print("  🚀 亚马逊 FBA 入库货件数据生成器（快速模式）")
    print_separator()
    print(f"  货件数量: {args.count}")
    print(f"  每个货件 SKU 数: {args.min_items} ~ {args.max_items}")
    print(f"  输出方式: {args.output.upper()}")

    generator = InboundDataGenerator(seed=args.seed)
    shipments, items = generator.generate_batch(
        count=args.count,
        shipment_config=shipment_config,
        item_config=item_config,
        min_items=args.min_items,
        max_items=args.max_items,
    )

    # 统计
    status_map = {
        1: "处理中", 2: "已发货", 3: "在途", 4: "已交付", 5: "检查中",
        6: "接收中", 7: "已完成", 8: "已取消", 9: "已删除", 10: "错误",
    }
    print(f"\n📊 货件状态分布:")
    status_count = {}
    for s in shipments:
        status = s.get("SHIPMENT_STATUS", 0)
        key = f"{status}-{status_map.get(status, '未知')}"
        status_count[key] = status_count.get(key, 0) + 1
    for k, v in sorted(status_count.items(), key=lambda x: int(x[0].split('-')[0])):
        bar = '█' * (v * 30 // max(args.count, 1))
        print(f"  {k:12s}: {v:4d}  {bar}")

    print(f"\n📊 数据统计:")
    print(f"  货件总数: {len(shipments)}")
    print(f"  明细总数: {len(items)}")

    # 输出
    if args.output == 'csv':
        export_to_csv(shipments, items, args.output_dir)
    elif args.output == 'sql':
        export_to_sql(shipments, items, args.sql_file)
    elif args.output == 'db':
        # 加载数据库配置（合并命令行参数和配置文件）
        db_cfg = load_db_config(getattr(args, 'db_config', None))
        # 命令行参数优先（仅当用户明确提供时覆盖）
        db_host = args.host if args._host_explicit else db_cfg["host"]
        db_port = args.port if args._port_explicit else db_cfg["port"]
        db_user = args.user if args._user_explicit else db_cfg["user"]
        db_pass = args.password if args._password_explicit else db_cfg["password"]
        db_name = args.database if args._database_explicit else db_cfg["database"]

        insert_to_mysql(
            shipments, items,
            host=db_host, port=db_port,
            user=db_user, password=db_pass,
            database=db_name,
            truncate=args.truncate,
            create_table=args.create_table,
            dry_run=args.dry_run,
        )

    print("\n🎉 完成！")


# ============================================================
# 入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description='亚马逊 FBA 入库货件数据生成脚本',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 交互式模式（推荐，可自定义字段）
  python generate_inbound_data.py

  # 快速模式（全部字段自动生成）
  python generate_inbound_data.py --count 50 --output csv

  # 使用已保存的模板
  python generate_inbound_data.py --template my_config
        """
    )

    # 交互式参数
    parser.add_argument('--interactive', '-i', action='store_true', default=True,
                        help='交互式模式（默认开启）')
    parser.add_argument('--template', type=str, default=None,
                        help='使用已保存的 JSON 配置模板名称')
    parser.add_argument('--template-excel', type=str, default=None,
                        help='使用 Excel 配置模板文件路径')
    parser.add_argument('--export-excel-template', action='store_true',
                        help='导出空白 Excel 配置模板')
    parser.add_argument('--list-templates', action='store_true',
                        help='列出所有已保存的模板')

    # 快速模式参数
    parser.add_argument('--count', type=int, default=50, help='货件数量（默认 50）')
    parser.add_argument('--min-items', type=int, default=1, help='每个货件最少 SKU 数')
    parser.add_argument('--max-items', type=int, default=5, help='每个货件最多 SKU 数')
    parser.add_argument('--seed', type=int, default=None, help='随机种子')
    parser.add_argument('--output', choices=['csv', 'db', 'sql'], default=None,
                        help='输出方式（指定后跳过交互式，直接快速生成）')
    parser.add_argument('--output-dir', type=str, default='./data')
    parser.add_argument('--sql-file', type=str, default='./data/insert_data.sql')
    # 数据库连接参数（命令行优先，其次 db_config.json，再其次环境变量）
    parser.add_argument('--host', type=str, default=None)
    parser.add_argument('--port', type=int, default=None)
    parser.add_argument('--user', type=str, default=None)
    parser.add_argument('--password', type=str, default=None)
    parser.add_argument('--database', type=str, default=None)
    parser.add_argument('--db-config', type=str, default=None,
                        help='数据库配置文件路径（默认: 脚本同目录/db_config.json）')

    # 数据库操作选项
    parser.add_argument('--truncate', action='store_true',
                        help='插入数据前清空目标表')
    parser.add_argument('--create-table', action='store_true',
                        help='自动创建目标表（IF NOT EXISTS）')
    parser.add_argument('--dry-run', action='store_true',
                        help='仅展示将要执行的 SQL 语句，不实际写入')
    parser.add_argument('--test-conn', action='store_true',
                        help='测试数据库连接是否可用')
    parser.add_argument('--export-ddl', action='store_true',
                        help='导出建表 DDL 到 SQL 文件')

    args = parser.parse_args()

    # 标记哪些数据库参数是用户明确指定的（用于优先级判断）
    args._host_explicit = args.host is not None
    args._port_explicit = args.port is not None
    args._user_explicit = args.user is not None
    args._password_explicit = args.password is not None
    args._database_explicit = args.database is not None

    # 填充默认值（后续会与 db_config.json 合并）
    args.host = args.host or '127.0.0.1'
    args.port = args.port or 3306
    args.user = args.user or 'root'
    args.password = args.password or ''
    args.database = args.database or 'test'

    # 测试数据库连接
    if args.test_conn:
        db_cfg = load_db_config(args.db_config)
        # 命令行参数覆盖
        host = args.host if args._host_explicit else db_cfg["host"]
        port = args.port if args._port_explicit else db_cfg["port"]
        user = args.user if args._user_explicit else db_cfg["user"]
        password = args.password if args._password_explicit else db_cfg["password"]
        database = args.database if args._database_explicit else db_cfg["database"]
        print_separator()
        print("  🔌 测试数据库连接")
        print_separator()
        print(f"  主机: {host}:{port}")
        print(f"  用户: {user}")
        print(f"  数据库: {database}")
        test_mysql_connection(host, port, user, password, database)
        return

    # 导出建表 DDL
    if args.export_ddl:
        ddl1 = generate_create_table_ddl("mws_fi_data_inbound_shipment", SHIPMENT_FIELDS)
        ddl2 = generate_create_table_ddl("mws_fi_data_inbound_shipment_item", ITEM_FIELDS)
        ddl_file = args.sql_file or './data/create_tables.sql'
        dir_path = os.path.dirname(ddl_file)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)
        with open(ddl_file, 'w', encoding='utf-8') as f:
            f.write("-- 亚马逊 FBA 入库货件表结构 DDL\n")
            f.write(f"-- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write("-- 货件主表\n")
            f.write(ddl1 + "\n\n")
            f.write("-- 货件明细表\n")
            f.write(ddl2 + "\n")
        print(f"✅ DDL 文件已导出: {ddl_file}")
        print(f"\n--- DDL 内容 ---")
        print(ddl1)
        print()
        print(ddl2)
        return

    # 列出模板
    if args.list_templates:
        templates = list_templates()
        if templates:
            print("已保存的配置模板:")
            for t in templates:
                icon = "📊" if t.get("format") == "excel" else "📄"
                fmt_label = "Excel" if t.get("format") == "excel" else "JSON"
                print(f"  {icon} {t['name']}  ({fmt_label})  修改于 {t['created_at']}")
        else:
            print("暂无已保存的模板。")
            print("  提示: 运行 --export-excel-template 可生成 Excel 模板到 excel_templates/ 目录")
        return

    # 导出 Excel 模板
    if args.export_excel_template:
        output_path = None  # 默认保存到 excel_templates/
        export_template_excel(output_path)
        return

    # 指定了 --output 则走快速模式
    # 使用 db 模式时，自动保存数据库配置（首次使用提示）
    if args.output == 'db' and not os.path.exists(DB_CONFIG_FILE) and not args.db_config:
        print(f"\n💡 提示: 可以通过 db_config.json 保存数据库连接配置，避免每次输入。")
        print(f"   配置文件路径: {DB_CONFIG_FILE}")
        print(f"   格式: {{\"host\": \"127.0.0.1\", \"port\": 3306, \"user\": \"root\", \"password\": \"xxx\", \"database\": \"test\"}}")
        if yes_no("  是否保存当前数据库配置?", "n"):
            db_cfg = load_db_config(args.db_config)
            save_db_config(db_cfg)

    if args.output:
        # 如果同时指定了 Excel 模板
        if args.template_excel:
            excel_path = _resolve_excel_template_path(args.template_excel)
            template = load_template_excel(excel_path)
            if not template:
                sys.exit(1)
            shipment_config = template.get("shipment", {})
            item_config = template.get("item", {})
        else:
            shipment_config = {fname: {"mode": "auto"} for fname in SHIPMENT_FIELDS}
            item_config = {fname: {"mode": "auto"} for fname in ITEM_FIELDS}

        cli_main_with_config(args, shipment_config, item_config)
        return

    # 否则走交互式
    if args.template_excel:
        # 使用 Excel 模板 + 交互式输入数量和输出方式
        excel_path = _resolve_excel_template_path(args.template_excel)
        template = load_template_excel(excel_path)
        if template:
            interactive_main_with_template(args, template)
        else:
            sys.exit(1)
    elif args.template:
        template = load_template(args.template)
        if template:
            print(f"✅ 已加载模板: {args.template}")
            # 用模板快速生成
            generator = InboundDataGenerator(seed=args.seed)
            shipments, items = generator.generate_batch(
                count=args.count,
                shipment_config=template.get("shipment", {}),
                item_config=template.get("item", {}),
                min_items=args.min_items,
                max_items=args.max_items,
            )
            print(f"📊 货件: {len(shipments)}, 明细: {len(items)}")
            export_to_csv(shipments, items, args.output_dir)
            print("\n🎉 完成！")
        else:
            print(f"⚠️  模板 '{args.template}' 不存在，进入交互式配置")
            interactive_main()
    else:
        interactive_main()


if __name__ == '__main__':
    main()
